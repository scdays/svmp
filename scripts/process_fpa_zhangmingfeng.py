#!/usr/bin/env python3
"""处理 FPA 中责任人为张明峰、J列修改说明为「新增」或「添加引用」的行（新版列布局）。"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

SRC = Path("/workspace/docs/fpa/FPA.xlsx")
EXPORT = Path("/workspace/docs/fpa/张明峰-功能点估算.csv")
OWNER = "张明峰"

# 新版列：I=类型 J=修改说明 V=FTR明细 W=RET明细 X=DET明细
COL = {"I": 9, "J": 10, "V": 22, "W": 23, "X": 24, "Y": 25}

UPDATES: dict[int, dict[str, str]] = {}
REF_UPDATES: dict[int, dict] = {}


def ilf_wx(table: str, fields: list[str]) -> dict[str, str]:
    assert 8 <= len(fields) <= 16, f"{table} field count {len(fields)}"
    return {"W": table, "X": f"{table}{{{('、'.join(fields))}}}"}


def trx(v_tables: list[str], det_parts: list[str]) -> dict[str, str]:
    return {"V": "、".join(v_tables), "X": "、".join(det_parts)}


def merge_v(existing: str | None, add: list[str]) -> str:
    parts: list[str] = []
    seen: set[str] = set()
    for chunk in (existing or "").split("、"):
        chunk = chunk.strip()
        if chunk and chunk not in seen:
            parts.append(chunk)
            seen.add(chunk)
    for name in add:
        if name not in seen:
            parts.append(name)
            seen.add(name)
    return "、".join(parts)


def merge_x(existing: str | None, add_parts: list[str]) -> str:
    if not existing:
        return "、".join(add_parts)
    parts = [p.strip() for p in existing.split("、") if p.strip()]
    seen = set(parts)
    for p in add_parts:
        if p not in seen:
            parts.append(p)
            seen.add(p)
    return "、".join(parts)


def build_updates() -> None:
    UPDATES.update(
        {
            110: ilf_wx(
                "漏洞组件表",
                [
                    "组件编号",
                    "漏洞编号",
                    "组件名称",
                    "组件版本",
                    "影响范围描述",
                    "关联资产编号",
                    "关联任务编号",
                    "发现时间",
                    "更新状态",
                    "数据来源",
                ],
            ),
            111: ilf_wx(
                "预警任务台账表",
                [
                    "台账编号",
                    "任务编号",
                    "任务名称",
                    "台账类型编码",
                    "记录时间",
                    "漏洞数量",
                    "资产数量",
                    "碰撞结果摘要",
                    "上报状态",
                    "部侧指令编号",
                ],
            ),
            112: ilf_wx(
                "产品漏洞预警工单表",
                [
                    "工单编号",
                    "工单名称",
                    "预警任务编号",
                    "工单状态",
                    "排查范围摘要",
                    "创建时间",
                    "下发时间",
                    "处理人",
                    "漏洞数量",
                    "资产数量",
                    "更新时间",
                ],
            ),
            197: ilf_wx(
                "插件参数表",
                [
                    "参数编号",
                    "插件编号",
                    "参数名称",
                    "参数类型",
                    "默认值",
                    "是否必填",
                    "参数说明",
                    "排序号",
                    "创建时间",
                    "更新时间",
                ],
            ),
            296: ilf_wx(
                "插件审核内容安全检测表",
                [
                    "检测记录编号",
                    "审核单编号",
                    "插件编号",
                    "插件版本号",
                    "检测时间",
                    "检测结论",
                    "检测报告摘要",
                    "检测引擎标识",
                    "检测人编号",
                    "创建时间",
                ],
            ),
            376: ilf_wx(
                "接口信息表",
                [
                    "接口编号",
                    "接口名称",
                    "接口地址",
                    "请求方式",
                    "接口版本",
                    "接口说明",
                    "启用状态",
                    "创建时间",
                    "更新时间",
                    "维护人",
                ],
            ),
            377: ilf_wx(
                "接口响应数据表",
                [
                    "响应编号",
                    "调用记录编号",
                    "响应状态码",
                    "响应头信息",
                    "响应体内容",
                    "响应时间",
                    "解析状态",
                    "异常说明",
                    "存储路径",
                    "创建时间",
                ],
            ),
            378: ilf_wx(
                "接口请求数据表",
                [
                    "请求编号",
                    "调用记录编号",
                    "请求头信息",
                    "请求体内容",
                    "请求时间",
                    "签名信息",
                    "重试次数",
                    "发送状态",
                    "失败原因",
                    "创建时间",
                ],
            ),
            79: trx(
                ["产品漏洞预警工单表"],
                [
                    "产品漏洞预警工单表{工单状态、预警任务编号|工单编号、工单名称、工单状态、漏洞数量、资产数量、创建时间、下发时间}"
                ],
            ),
            80: trx(
                ["产品漏洞预警工单表"],
                [
                    "产品漏洞预警工单表{工单编号|工单名称、工单状态、排查范围摘要、漏洞数量、资产数量、处理人、创建时间、下发时间}"
                ],
            ),
            81: trx(
                ["产品漏洞预警工单表", "扫描资产表", "产品漏洞表"],
                [
                    "产品漏洞预警工单表{工单编号、排查范围摘要|工单编号、排查范围摘要、更新时间}",
                    "扫描资产表{资产编号、主机名、操作系统|资产编号、主机名、操作系统、部门}",
                    "产品漏洞表{漏洞编号、漏洞名称|漏洞编号、漏洞名称、影响程度}",
                ],
            ),
            82: trx(
                ["系统漏洞结果表", "产品漏洞表"],
                [
                    "系统漏洞结果表{任务实例编号|结果编号、漏洞编号、资产编号、检测结果、修复建议}",
                    "产品漏洞表{漏洞编号|漏洞编号、漏洞名称、影响程度、发现时间}",
                ],
            ),
            83: trx(
                ["扫描资产表"],
                ["扫描资产表{任务编号、部门、操作系统|资产编号、主机名、操作系统、部门、IP地址}"],
            ),
            84: trx(
                ["产品漏洞表"],
                ["产品漏洞表{漏洞名称、影响程度|漏洞编号、漏洞名称、描述、影响程度、发现时间}"],
            ),
            89: trx(
                ["产品漏洞预警任务表"],
                [
                    "产品漏洞预警任务表{任务状态、开始时间|任务编号、任务名称、任务描述、开始时间、结束时间、状态}"
                ],
            ),
            90: trx(
                ["预警任务台账表"],
                [
                    "预警任务台账表{任务编号、台账类型编码、记录时间|台账编号、任务名称、漏洞数量、资产数量、上报状态}"
                ],
            ),
            91: trx(
                ["预警任务台账表"],
                [
                    "预警任务台账表{台账编号|台账编号、任务编号、任务名称、记录时间、漏洞数量、资产数量、碰撞结果摘要、上报状态}"
                ],
            ),
            184: trx(
                ["插件版本管理表", "插件基本信息表"],
                [
                    "插件版本管理表{插件编号、版本号、发布说明、发布时间|版本编号、插件编号、版本号、发布说明、发布时间}",
                    "插件基本信息表{插件编号|插件编号、插件名称、版本、更新时间}",
                ],
            ),
            185: trx(
                ["插件版本管理表"],
                [
                    "插件版本管理表{版本编号、版本号、发布说明、发布时间|版本编号、插件编号、版本号、发布说明、发布时间}"
                ],
            ),
            186: trx(
                ["插件版本管理表"],
                ["插件版本管理表{版本编号|版本编号、插件编号}"],
            ),
            187: trx(
                ["插件版本管理表", "插件基本信息表"],
                [
                    "插件版本管理表{插件编号、版本号|版本编号、插件编号、版本号、发布说明、发布时间}",
                    "插件基本信息表{插件编号、插件名称|插件编号、插件名称、编码、状态}",
                ],
            ),
            193: trx(
                ["插件模板表", "插件规范表"],
                [
                    "插件模板表{模板名称、状态|模板编号、模板名称、模板内容、创建时间、状态}",
                    "插件规范表{规范名称、规范版本|规范编号、规范名称、规范版本、规范内容、创建时间}",
                ],
            ),
            199: trx(
                ["插件分类表"],
                [
                    "插件分类表{分类编号、分类名称、父分类编号、排序号、说明|分类编号、分类名称、父分类编号、排序号、说明}"
                ],
            ),
            221: trx(
                ["原理检测插件审核单表", "插件审核内容安全检测表"],
                [
                    "原理检测插件审核单表{审核单编号、插件编号、插件版本号|审核单编号、审核单状态、更新时间}",
                    "插件审核内容安全检测表{审核单编号、插件编号、插件版本号、检测结论|检测记录编号、检测时间、检测结论、检测报告摘要}",
                ],
            ),
            222: trx(
                ["原理检测插件审核单表", "插件编译表"],
                [
                    "原理检测插件审核单表{审核单编号|审核单编号、审核单状态}",
                    "插件编译表{审核单编号、插件编号、插件版本号、编译结论、运行结论|编译运行记录编号、执行时间、编译结论、运行结论、日志摘要}",
                ],
            ),
            223: trx(
                ["原理检测插件审核单表", "插件综合判定表"],
                [
                    "原理检测插件审核单表{审核单编号|审核单编号、审核单状态}",
                    "插件综合判定表{审核单编号、插件编号、插件版本号、综合状态、判定说明|判定记录编号、综合状态、判定时间、判定说明}",
                ],
            ),
        }
    )

    REF_UPDATES.update(
        {
            137: {
                "add_v": ["纠偏资产表", "任务执行实例表"],
                "add_x": [
                    "纠偏资产表{任务编号、资产编号、原始名称|资产编号、任务编号、原始名称、对齐后名称、状态}",
                    "任务执行实例表{任务编号、资产编号、状态|实例编号、任务编号、资产编号、状态、开始时间、结束时间}",
                ],
            },
            200: {
                "add_v": ["插件分类表"],
                "add_x": [
                    "插件分类表{分类编号、分类名称、父分类编号|分类编号、分类名称、父分类编号、排序号、说明}"
                ],
            },
            201: {
                "add_v": ["插件分类表"],
                "add_x": ["插件分类表{分类编号|分类编号、分类名称、父分类编号、排序号、说明}"],
            },
            314: {
                "add_v": ["插件部署状态表", "安全资源注册请求参数表"],
                "add_x": [
                    "插件部署状态表{插件编号、插件版本号、部署环境、部署状态|部署记录编号、部署时间}",
                    "安全资源注册请求参数表{插件编号、安全资源描述、IP范围|请求编号}",
                ],
            },
            315: {
                "add_v": ["插件部署状态表", "插件执行记录表"],
                "add_x": [
                    "插件部署状态表{插件编号、部署环境、部署状态|部署记录编号、部署状态、部署时间}",
                    "插件执行记录表{插件编号、实例编号、执行结果|执行记录编号、开始时间、结束时间、执行结果、日志摘要}",
                ],
            },
            316: {
                "add_v": ["插件部署状态表", "插件执行记录表"],
                "add_x": [
                    "插件部署状态表{插件编号|部署记录编号、部署状态}",
                    "插件执行记录表{插件编号|执行记录编号、开始时间、结束时间、执行结果}",
                ],
            },
            335: {
                "add_v": ["插件任务高级配置表", "插件任务基本信息表"],
                "add_x": [
                    "插件任务高级配置表{任务编号、配置项键、配置项值|配置编号、更新时间}",
                    "插件任务基本信息表{任务编号|任务编号、插件编号、任务名称、任务状态}",
                ],
            },
            346: {
                "add_v": ["报告生成记录表"],
                "add_x": ["报告生成记录表{报告编号、任务编号|记录编号、生成时间、生成人}"],
            },
        }
    )


def apply_static_updates(ws) -> None:
    for row, cols in UPDATES.items():
        for key, value in cols.items():
            ws.cell(row, COL[key]).value = value


def apply_reference_updates(ws) -> None:
    for row, spec in REF_UPDATES.items():
        old_v = ws.cell(row, COL["V"]).value
        old_x = ws.cell(row, COL["X"]).value
        ws.cell(row, COL["V"]).value = merge_v(old_v, spec["add_v"])
        ws.cell(row, COL["X"]).value = merge_x(old_x, spec["add_x"])


def export_zhangmingfeng(ws, out: Path) -> int:
    headers = [ws.cell(1, c).value for c in range(1, 30)]
    rows_out = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 4).value != OWNER:
            continue
        rows_out.append([ws.cell(r, c).value for c in range(1, 30)])
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows_out)
    return len(rows_out)


def main() -> None:
    build_updates()
    wb = openpyxl.load_workbook(SRC)
    ws = wb["功能点估算"]
    apply_static_updates(ws)
    apply_reference_updates(ws)
    n = export_zhangmingfeng(ws, EXPORT)
    wb.save(SRC)
    print(f"static={len(UPDATES)} ref={len(REF_UPDATES)} exported={n}")


if __name__ == "__main__":
    main()
