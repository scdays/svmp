#!/usr/bin/env python3
"""Fill FPA.xlsx rows 727-856 for 漏管台账管理 per spec tables 53-57."""

import openpyxl

SRC = "/workspace/docs/fpa/FPA.xlsx"
START, END = 727, 856
FORMULA_COLS = ["A", "E", "M", "N", "O", "P", "Q", "S", "U", "V", "Z", "AA"]

LOG_MAIN = "漏管台账日志主表{logID、logType、logLvl、timeStamp、devHash、loginAccount、loginIp、orderID、l2Code、bkItemID、dataFileID、chainHash、chainHashR、content}"
LOG_CONTENT = "台账日志content扩展表{logID、logType、engHash、tskName、tskId、target、targetPort、targetFileLoc、targetPortFileLoc、tskProgress、vulNum、vulInfoID、vulInfoStat、vulReqMsg、seqNum、opRslt、srcTktRole、srcTktPrsn、dstTktRole、dstTktPrsn、tktId、tktName、method、remedDesc、remedTime、remedInfo、astNum、vulAstNum、vulInfoNum}"
CHAIN = "日志链式哈希索引表{logID、chainHash、chainHashR、prevLogID、hashTime}"
FILE_META = "日志配套文件元数据表{dataFileID、orderID、fileKey、filePath、fileSize、createTime}"
LOGTYPE_CFG = "logType码表配置表{logType、logTypeName、logLvlRule、contentSchema、enabled}"
MAINT = "漏管平台维护日志表{logID、logType、logLvl、timeStamp、loginAccount、loginIp、devHash、operAction、operDesc}"
SYSLOG = "漏管平台系统日志表{logID、logType、logLvl、timeStamp、syncStatus、syncSource、devHash}"
SCAN = "系统漏洞排查弱口令扫描日志表{logID、logType、engHash、tskName、tskId、target、targetPort、targetFileLoc、targetPortFileLoc、tskProgress、vulNum}"
EXPLOIT = "系统漏洞利用日志表{logID、logType、vulInfoID、vulInfoStat、engHash、vulReqMsg、seqNum、opRslt}"
FIX = "系统漏洞修复日志表{logID、logType、tktId、tktName、srcTktRole、srcTktPrsn、dstTktRole、dstTktPrsn、method、remedDesc、remedTime、remedInfo}"
WARN = "产品漏洞预警日志表{logID、logType、tskName、tskId、vulNum、astNum、vulAstNum、vulInfoNum}"
FIX_REL = "修复工单台账关联表{tktId、logID、bkItemID、orderID、linkTime}"
COLLISION = "产品漏洞碰撞分析结果表{logID、vulInfoID、assetID、collisionResult、analysisTime}"
REPORT_CFG = "台账上报周期配置表{configID、cronExpr、reportScope、enabled、updateTime}"
REPORT_CONTENT = "台账上报内容表{batchID、logReqSeq、numLogs、reportStatus、reportTime、failReason}"
REPORT_HIST = "台账日志上报历史表{historyID、batchID、reportType、reportStatus、reportTime、responseCode}"
REPORT_MON = "台账上报状态监控表{monitorID、pendingCount、successCount、failCount、lastReportTime}"
MINISTRY_REQ = "部侧logInfoReqParams上报接口文件{fileID、orderID、logReqSeq、numLogs、filePath、uploadTime}"
MINISTRY_RSP = "部侧logInfoReqParams响应接口文件{fileID、orderID、responseCode、responseMsg、responseTime}"
SCENE11 = "业务场景11日志文件表{dataFileID、orderID、fileKey、logCount、filePath、createTime}"
TASSTA = "tasStaObj统计表{statID、statPeriod、logType、orderType、sendCount、recvCount}"
TEST_QUERY = "部侧测试日志查询记录表{queryID、testType、queryTime、responseStatus、logCount}"
VULMSG = "系统漏洞利用报文表{logID、seqNum、vulReqMsg、createTime}"
DEVICE = "注册扫描设备信息表{engHash、deviceName、deviceType、registerTime}"
LOGTYPE_STAT = "logType台账数量统计表{statID、logType、recordCount、statPeriod}"
REPORT_TREND = "台账上报趋势统计表{statID、statDate、successCount、failCount}"
ORDER_STAT = "指令类型收发统计表{statID、orderType、sendCount、recvCount、statPeriod}"


def build_rows():
    rows = []

    rows += [
        dict(f="漏管台账管理", g="台账日志基础能力", h="台账日志数据存储", i="漏管台账日志主表", j="ILF", x="漏管台账日志主表", y=LOG_MAIN, z="表53"),
        dict(h="台账日志数据存储", i="台账日志content扩展表", j="ILF", x="台账日志content扩展表", y=LOG_CONTENT, z="表54-57"),
        dict(h="台账日志数据存储", i="日志链式哈希索引表", j="ILF", x="日志链式哈希索引表", y=CHAIN, z="表53 chainHash"),
        dict(h="台账日志数据存储", i="日志配套文件元数据表", j="ILF", x="日志配套文件元数据表", y=FILE_META, z="业务场景11"),
        dict(h="配置管理", i="logType码表配置表", j="ILF", x="logType码表配置表", y=LOGTYPE_CFG, z="附录A.10"),
        dict(h="部侧接口对接", i="部侧logInfoReqParams上报接口文件", j="EIF", x="部侧logInfoReqParams上报接口文件", y=MINISTRY_REQ, ab="EIF"),
        dict(h="部侧接口对接", i="部侧logInfoReqParams响应接口文件", j="EIF", x="部侧logInfoReqParams响应接口文件", y=MINISTRY_RSP, ab="EIF"),
        dict(h="日志采集写入", i="接收业务事件写入logInfo", j="EI", w="漏管台账日志主表、台账日志content扩展表", y="漏管台账日志主表{logID、logType、logLvl、timeStamp、devHash、loginAccount、loginIp|logID}\n台账日志content扩展表{logID、content|logID、content}"),
        dict(h="日志采集写入", i="组装logInfoReqParams批量头", j="EI", w="漏管台账日志主表", y="漏管台账日志主表{logReqSeq、numLogs、timeStamp|batchID、logID}"),
        dict(h="日志采集写入", i="关联上级驱动字段orderID与bkItemID", j="EI", w="漏管台账日志主表", y="漏管台账日志主表{orderID、l2Code、bkItemID|logID、orderID、l2Code、bkItemID}"),
        dict(h="日志采集写入", i="计算并写入chainHash链式哈希", j="EI", w="漏管台账日志主表、日志链式哈希索引表", y="漏管台账日志主表{logID、chainHash|chainHash、chainHashR}\n日志链式哈希索引表{logID、prevLogID|chainHash、hashTime}"),
        dict(h="日志采集写入", i="工单回传时携带logInfoReqParams", j="EO", w="漏管台账日志主表、台账上报内容表", y="漏管台账日志主表{orderID、bkItemID|logID、logType}\n台账上报内容表{|batchID、numLogs}"),
        dict(h="日志采集写入", i="常态化上报未关联工单的系统日志", j="EO", w="漏管台账日志主表", y="漏管台账日志主表{logType、timeStamp|logID、logType、logLvl}"),
        dict(h="日志上报输出", i="向部侧上报logInfoReqParams", j="EO", w="漏管台账日志主表、部侧logInfoReqParams上报接口文件", y="漏管台账日志主表{logID、logType|reportStatus}\n部侧logInfoReqParams上报接口文件{|fileID、filePath、uploadTime}"),
        dict(h="日志查询", i="按logID或bkItemID或orderID查询台账日志", j="EQ", w="漏管台账日志主表、台账日志content扩展表", y="漏管台账日志主表{logID、orderID、bkItemID|logID、logType、logLvl、timeStamp}\n台账日志content扩展表{logID|content}"),
        dict(h="日志查询", i="按logType筛选台账日志", j="EQ", w="漏管台账日志主表", y="漏管台账日志主表{logType、timeStamp|logID、logType、logLvl、orderID}"),
        dict(h="日志查询", i="按时间范围筛选台账日志", j="EQ", w="漏管台账日志主表", y="漏管台账日志主表{timeStamp、logType|logID、logType、timeStamp}"),
        dict(h="日志查询", i="导出台账日志查询结果", j="EO", w="漏管台账日志主表、台账日志content扩展表", y="漏管台账日志主表{logID、logType|导出文件路径}"),
        dict(h="日志查询", i="响应部侧测试类日志查询请求", j="EO", w="漏管台账日志主表、部侧测试日志查询记录表", y="漏管台账日志主表{logType、timeStamp|logID、logType、content}\n部侧测试日志查询记录表{testType|queryID、responseStatus、logCount}", z="测试3/6"),
        dict(h="日志查询", i="校验台账日志链式哈希完整性", j="EQ", w="日志链式哈希索引表、漏管台账日志主表", y="日志链式哈希索引表{logID|chainHash、chainHashR、prevLogID}"),
        dict(h="日志配套与统计", i="打包业务场景11日志配套文件", j="EO", w="业务场景11日志文件表、漏管台账日志主表", y="漏管台账日志主表{logID、dataFileID|dataFileID}\n业务场景11日志文件表{|dataFileID、fileKey、logCount、filePath}"),
        dict(h="日志配套与统计", i="维护dataFileID与orderID文件传输关联", j="EI", w="日志配套文件元数据表", y="日志配套文件元数据表{dataFileID、orderID、fileKey|dataFileID、orderID、filePath}"),
        dict(h="日志配套与统计", i="下载业务场景11日志配套文件", j="EO", w="业务场景11日志文件表", y="业务场景11日志文件表{dataFileID|filePath、fileKey}"),
        dict(h="日志配套与统计", i="查询业务场景11配套文件列表", j="EQ", w="业务场景11日志文件表", y="业务场景11日志文件表{orderID、dataFileID|fileKey、logCount、createTime}"),
    ]

    rows += [
        dict(g="漏管平台系统日志", h="平台维护日志1010", i="采集登录与管理操作生成1010日志", j="EI", w="漏管平台维护日志表", y="漏管平台维护日志表{loginAccount、loginIp、devHash、operAction|logID、logType、timeStamp}"),
        dict(h="平台维护日志1010", i="定时采集时钟同步事件生成1000日志", j="EI", w="漏管平台系统日志表", y="漏管平台系统日志表{syncStatus、syncSource、devHash|logID、logType、logLvl、timeStamp}", z="每10分钟"),
        dict(h="系统运行日志1000", i="生成logType1000系统日志并纳入上报批次", j="EO", w="漏管平台系统日志表、漏管台账日志主表", y="漏管平台系统日志表{logID|reportStatus}\n漏管台账日志主表{|logID、logType、logLvl}"),
        dict(h="平台维护日志1010", i="查询漏管平台维护日志", j="EQ", w="漏管平台维护日志表", y="漏管平台维护日志表{logID、timeStamp、loginAccount|logID、logType、loginAccount、operAction}"),
        dict(h="平台维护日志1010", i="查看漏管平台维护日志摘要", j="EQ", w="漏管平台维护日志表", y="漏管平台维护日志表{logID|logType、timeStamp、operAction、operDesc}"),
        dict(h="平台维护日志1010", i="导出漏管平台维护日志文件", j="EO", w="漏管平台维护日志表", y="漏管平台维护日志表{logID、timeStamp|导出文件路径}"),
        dict(h="平台维护日志1010", i="漏管平台维护日志表", j="ILF", x="漏管平台维护日志表", y=MAINT, z="logType=1010"),
        dict(h="系统运行日志1000", i="查询漏管平台系统日志", j="EQ", w="漏管平台系统日志表", y="漏管平台系统日志表{logID、timeStamp|logID、logType、logLvl、syncStatus}"),
        dict(h="系统运行日志1000", i="查看漏管平台系统日志摘要", j="EQ", w="漏管平台系统日志表", y="漏管平台系统日志表{logID|logType、timeStamp、syncStatus}"),
        dict(h="系统运行日志1000", i="导出漏管平台系统日志文件", j="EO", w="漏管平台系统日志表", y="漏管平台系统日志表{logID、timeStamp|导出文件路径}"),
        dict(h="系统运行日志1000", i="漏管平台系统日志表", j="ILF", x="漏管平台系统日志表", y=SYSLOG, z="logType=1000"),
    ]

    rows += [
        dict(g="系统漏洞排查日志", h="排查任务日志1020", i="记录排查任务下发审计日志", j="EI", w="系统漏洞排查弱口令扫描日志表", y="系统漏洞排查弱口令扫描日志表{engHash、tskName、tskId、target|logID、logType、timeStamp}", z="表54"),
        dict(h="排查任务日志1020", i="记录排查任务进度审计日志", j="EI", w="系统漏洞排查弱口令扫描日志表", y="系统漏洞排查弱口令扫描日志表{tskProgress、targetPort|logID、tskProgress}"),
        dict(h="排查任务日志1020", i="记录排查任务结果审计日志", j="EI", w="系统漏洞排查弱口令扫描日志表", y="系统漏洞排查弱口令扫描日志表{vulNum、tskProgress|logID、vulNum}"),
        dict(h="排查任务日志1020", i="生成target与targetPort配套数据文件", j="EO", w="系统漏洞排查弱口令扫描日志表、日志配套文件元数据表", y="系统漏洞排查弱口令扫描日志表{target、targetPort|targetFileLoc、targetPortFileLoc}\n日志配套文件元数据表{|dataFileID、fileKey}"),
        dict(h="弱口令扫描日志1030", i="记录弱口令扫描任务台账日志", j="EI", w="系统漏洞排查弱口令扫描日志表", y="系统漏洞排查弱口令扫描日志表{engHash、tskName、tskId|logID、logType=1030}"),
        dict(h="设备关联", i="注册扫描设备信息表", j="EIF", x="注册扫描设备信息表", y=DEVICE, ab="EIF"),
        dict(h="排查任务日志1020", i="查询系统漏洞排查弱口令扫描日志", j="EQ", w="系统漏洞排查弱口令扫描日志表", y="系统漏洞排查弱口令扫描日志表{logID、logType、tskId|logID、engHash、tskName、tskProgress、vulNum}"),
        dict(h="排查任务日志1020", i="查看排查弱口令扫描日志摘要", j="EQ", w="系统漏洞排查弱口令扫描日志表", y="系统漏洞排查弱口令扫描日志表{logID|logType、tskName、tskProgress、vulNum}"),
        dict(h="排查任务日志1020", i="导出排查弱口令扫描日志文件", j="EO", w="系统漏洞排查弱口令扫描日志表", y="系统漏洞排查弱口令扫描日志表{logID、logType|导出文件路径}"),
        dict(h="排查任务日志1020", i="系统漏洞排查弱口令扫描日志表", j="ILF", x="系统漏洞排查弱口令扫描日志表", y=SCAN, z="1020/1030"),
    ]

    rows += [
        dict(g="系统漏洞利用日志", h="利用过程日志1040", i="记录EXP利用过程日志", j="EI", w="系统漏洞利用日志表", y="系统漏洞利用日志表{vulInfoID、vulInfoStat、engHash、seqNum|logID、logType、timeStamp}", z="表55"),
        dict(h="利用过程日志1040", i="系统漏洞利用报文表", j="ILF", x="系统漏洞利用报文表", y=VULMSG),
        dict(h="利用过程日志1040", i="写入vulReqMsg交互报文", j="EI", w="系统漏洞利用报文表", y="系统漏洞利用报文表{logID、seqNum、vulReqMsg|logID、seqNum}"),
        dict(h="利用过程日志1040", i="校验利用结果opRslt", j="EI", w="系统漏洞利用日志表", y="系统漏洞利用日志表{opRslt、vulInfoID|logID、opRslt}"),
        dict(h="利用过程日志1040", i="查询系统漏洞利用日志", j="EQ", w="系统漏洞利用日志表", y="系统漏洞利用日志表{logID、vulInfoID|logID、logType、engHash、opRslt}"),
        dict(h="利用过程日志1040", i="查看漏洞利用日志摘要", j="EQ", w="系统漏洞利用日志表", y="系统漏洞利用日志表{logID|vulInfoID、opRslt}"),
        dict(h="利用过程日志1040", i="导出系统漏洞利用日志文件", j="EO", w="系统漏洞利用日志表", y="系统漏洞利用日志表{logID|导出文件路径}"),
        dict(h="利用过程日志1040", i="系统漏洞利用日志表", j="ILF", x="系统漏洞利用日志表", y=EXPLOIT),
    ]

    rows += [
        dict(g="系统漏洞修复日志", h="修复工单日志1050", i="记录修复工单指派审计日志", j="EI", w="系统漏洞修复日志表", y="系统漏洞修复日志表{srcTktRole、srcTktPrsn、dstTktRole、dstTktPrsn、tktId、tktName|logID、logType}", z="表56"),
        dict(h="修复工单日志1050", i="记录修复完成上报审计日志", j="EI", w="系统漏洞修复日志表", y="系统漏洞修复日志表{method、remedDesc、remedTime、remedInfo|logID、method、remedTime}"),
        dict(h="修复工单日志1050", i="组装remedInfo漏洞修复列表", j="EI", w="系统漏洞修复日志表", y="系统漏洞修复日志表{remedInfo、vulInfoID、vulInfoStat|logID、remedInfo}"),
        dict(h="修复工单日志1050", i="维护修复工单与台账日志关联", j="EI", w="修复工单台账关联表", y="修复工单台账关联表{tktId、logID、bkItemID、orderID|linkTime}"),
        dict(h="修复工单日志1050", i="修复工单台账关联表", j="ILF", x="修复工单台账关联表", y=FIX_REL),
        dict(h="修复工单日志1050", i="查询系统漏洞修复日志", j="EQ", w="系统漏洞修复日志表", y="系统漏洞修复日志表{logID、tktId|logID、logType、tktName、method}"),
        dict(h="修复工单日志1050", i="查看修复工单日志摘要", j="EQ", w="系统漏洞修复日志表", y="系统漏洞修复日志表{logID|tktName、method、remedDesc}"),
        dict(h="修复工单日志1050", i="导出系统漏洞修复日志文件", j="EO", w="系统漏洞修复日志表", y="系统漏洞修复日志表{logID|导出文件路径}"),
        dict(h="修复工单日志1050", i="系统漏洞修复日志表", j="ILF", x="系统漏洞修复日志表", y=FIX),
    ]

    rows += [
        dict(g="产品漏洞预警日志", h="预警碰撞日志1080", i="创建产品漏洞预警分析任务日志", j="EI", w="产品漏洞预警日志表", y="产品漏洞预警日志表{tskName、tskId|logID、logType、timeStamp}", z="表57"),
        dict(h="预警碰撞日志1080", i="记录下发表单统计vulNum与astNum", j="EI", w="产品漏洞预警日志表", y="产品漏洞预警日志表{vulNum、astNum|logID、vulNum、astNum}"),
        dict(h="预警碰撞日志1080", i="记录碰撞结果vulAstNum与vulInfoNum", j="EI", w="产品漏洞预警日志表、产品漏洞碰撞分析结果表", y="产品漏洞预警日志表{vulAstNum、vulInfoNum|logID}\n产品漏洞碰撞分析结果表{vulInfoID、assetID|collisionResult}"),
        dict(h="预警碰撞日志1080", i="产品漏洞碰撞分析结果表", j="ILF", x="产品漏洞碰撞分析结果表", y=COLLISION),
        dict(h="预警碰撞日志1080", i="查询产品漏洞预警日志", j="EQ", w="产品漏洞预警日志表", y="产品漏洞预警日志表{logID、tskId|logID、tskName、vulNum、astNum}"),
        dict(h="预警碰撞日志1080", i="查看预警碰撞日志摘要", j="EQ", w="产品漏洞预警日志表", y="产品漏洞预警日志表{logID|tskName、vulAstNum、vulInfoNum}"),
        dict(h="预警碰撞日志1080", i="导出产品漏洞预警日志文件", j="EO", w="产品漏洞预警日志表", y="产品漏洞预警日志表{logID|导出文件路径}"),
        dict(h="预警碰撞日志1080", i="产品漏洞预警日志表", j="ILF", x="产品漏洞预警日志表", y=WARN),
    ]

    rows += [
        dict(g="台账数据上报管理", h="台账数据定期上报", i="配置台账日志上报周期", j="EI", w="台账上报周期配置表", y="台账上报周期配置表{cronExpr、reportScope、enabled|configID}"),
        dict(h="台账数据定期上报", i="定时组装并上报logInfoReqParams", j="EO", w="漏管台账日志主表、台账上报内容表、部侧logInfoReqParams上报接口文件", y="漏管台账日志主表{logType|logID}\n台账上报内容表{|batchID、reportStatus}\n部侧logInfoReqParams上报接口文件{|fileID}"),
        dict(h="台账数据定期上报", i="查询台账定时上报记录", j="EQ", w="台账日志上报历史表", y="台账日志上报历史表{reportType、reportTime|historyID、reportStatus}"),
        dict(h="台账数据手动上报", i="查询未成功上报的台账日志", j="EQ", w="台账上报内容表", y="台账上报内容表{reportStatus|batchID、failReason}"),
        dict(h="台账数据手动上报", i="预览待上报logInfoReqParams批次", j="EQ", w="台账上报内容表、漏管台账日志主表", y="台账上报内容表{batchID|logReqSeq、numLogs}\n漏管台账日志主表{batchID|logID、logType}"),
        dict(h="台账数据手动上报", i="人工主动上报台账日志批次", j="EO", w="台账上报内容表、部侧logInfoReqParams上报接口文件", y="台账上报内容表{batchID|reportStatus}\n部侧logInfoReqParams上报接口文件{|fileID}"),
        dict(h="台账数据手动上报", i="查询台账手动上报记录", j="EQ", w="台账日志上报历史表", y="台账日志上报历史表{reportType|historyID、reportStatus}"),
        dict(h="台账数据上报监控", i="台账上报状态监控", j="EO", w="台账上报状态监控表、台账日志上报历史表", y="台账上报状态监控表{|pendingCount、successCount}\n台账日志上报历史表{reportTime|reportStatus}"),
        dict(h="台账数据上报监控", i="查询台账上报历史", j="EQ", w="台账日志上报历史表", y="台账日志上报历史表{reportTime、reportStatus|historyID、batchID}"),
        dict(h="台账数据上报监控", i="导出台账上报历史", j="EO", w="台账日志上报历史表", y="台账日志上报历史表{reportTime|导出文件路径}"),
        dict(h="台账数据上报监控", i="台账上报失败告警通知", j="EO", w="台账上报状态监控表", y="台账上报状态监控表{failCount|monitorID、lastReportTime}"),
        dict(h="部侧接口对接", i="重试失败的上报批次", j="EI", w="台账上报内容表、部侧logInfoReqParams上报接口文件", y="台账上报内容表{batchID|reportStatus、failReason}"),
        dict(h="部侧接口对接", i="解析部侧logInfoReqParams响应", j="EI", w="部侧logInfoReqParams响应接口文件、台账日志上报历史表", y="部侧logInfoReqParams响应接口文件{fileID|responseCode、responseMsg}\n台账日志上报历史表{batchID|responseCode}"),
        dict(h="台账数据定期上报", i="台账上报周期配置表", j="ILF", x="台账上报周期配置表", y=REPORT_CFG),
        dict(h="台账数据定期上报", i="台账上报内容表", j="ILF", x="台账上报内容表", y=REPORT_CONTENT),
        dict(h="台账数据上报监控", i="台账日志上报历史表", j="ILF", x="台账日志上报历史表", y=REPORT_HIST),
        dict(h="台账数据上报监控", i="台账上报状态监控表", j="ILF", x="台账上报状态监控表", y=REPORT_MON),
        dict(h="日志配套与统计", i="业务场景11日志文件表", j="ILF", x="业务场景11日志文件表", y=SCENE11),
        dict(h="日志配套与统计", i="tasStaObj统计表", j="ILF", x="tasStaObj统计表", y=TASSTA, z="表51"),
        dict(h="日志查询", i="部侧测试日志查询记录表", j="ILF", x="部侧测试日志查询记录表", y=TEST_QUERY),
        dict(h="日志配套与统计", i="按logType统计tasStaObj周期台账数量", j="EO", w="tasStaObj统计表、漏管台账日志主表", y="漏管台账日志主表{logType、timeStamp|logType}\ntasStaObj统计表{|statPeriod、logType、sendCount}"),
        dict(h="日志配套与统计", i="按指令类型统计tasStaObj本地收发数量", j="EO", w="tasStaObj统计表", y="tasStaObj统计表{statPeriod、orderType|sendCount、recvCount}"),
    ]

    rows += [
        dict(g="台账数据可视化展示", h="logType分布统计", i="logType台账数量统计", j="EO", w="logType台账数量统计表、漏管台账日志主表", y="漏管台账日志主表{logType|logType}\nlogType台账数量统计表{|recordCount}"),
        dict(h="logType分布统计", i="logType台账数量统计表", j="ILF", x="logType台账数量统计表", y=LOGTYPE_STAT),
        dict(h="logType分布统计", i="台账数据类型分布图", j="EO", w="logType台账数量统计表", y="logType台账数量统计表{statPeriod|logType、recordCount}"),
        dict(h="上报趋势统计", i="台账上报趋势统计表", j="ILF", x="台账上报趋势统计表", y=REPORT_TREND),
        dict(h="上报趋势统计", i="上报执行趋势图", j="EO", w="台账上报趋势统计表、台账日志上报历史表", y="台账日志上报历史表{reportTime|reportStatus}\n台账上报趋势统计表{|successCount、failCount}"),
        dict(h="上报趋势统计", i="台账上报状态分布图", j="EO", w="台账上报状态监控表", y="台账上报状态监控表{|pendingCount、successCount、failCount}"),
        dict(h="指令收发统计", i="指令类型收发统计", j="EO", w="指令类型收发统计表、tasStaObj统计表", y="tasStaObj统计表{orderType|sendCount、recvCount}\n指令类型收发统计表{|statPeriod、orderType}"),
        dict(h="指令收发统计", i="指令类型收发统计表", j="ILF", x="指令类型收发统计表", y=ORDER_STAT),
        dict(g="台账日志基础能力", h="配置管理", i="维护logType码表映射规则", j="EI", w="logType码表配置表", y="logType码表配置表{logType、logTypeName、logLvlRule|enabled}"),
        dict(h="配置管理", i="查询logType码表配置", j="EQ", w="logType码表配置表", y="logType码表配置表{logType|logTypeName、logLvlRule、enabled}"),
        dict(h="配置管理", i="启用或停用logType采集规则", j="EI", w="logType码表配置表", y="logType码表配置表{logType、enabled|enabled}"),
        dict(h="日志查询", i="按bkItemID追溯关联工单日志", j="EQ", w="漏管台账日志主表", y="漏管台账日志主表{bkItemID|logID、logType、orderID、timeStamp}"),
        dict(h="日志查询", i="按orderID追溯部侧指令日志", j="EQ", w="漏管台账日志主表", y="漏管台账日志主表{orderID|logID、logType、bkItemID、timeStamp}"),
        dict(h="日志查询", i="按devHash追溯设备相关日志", j="EQ", w="漏管台账日志主表、系统漏洞排查弱口令扫描日志表", y="漏管台账日志主表{devHash|logID、logType}\n系统漏洞排查弱口令扫描日志表{engHash|logID、tskName}"),
        dict(h="部侧接口对接", i="模拟部侧测试类型3日志查询", j="EO", w="漏管台账日志主表、部侧测试日志查询记录表", y="漏管台账日志主表{logType|logID}\n部侧测试日志查询记录表{testType=3|queryID、logCount}", z="测试类型3"),
        dict(h="部侧接口对接", i="模拟部侧测试类型6日志查询", j="EO", w="漏管台账日志主表、部侧测试日志查询记录表", y="漏管台账日志主表{logType|logID}\n部侧测试日志查询记录表{testType=6|queryID、logCount}", z="测试类型6"),
        dict(h="日志配套与统计", i="生成tasStaObj统计快照", j="EO", w="tasStaObj统计表、漏管台账日志主表", y="漏管台账日志主表{logType、timeStamp|logType}\ntasStaObj统计表{|statID、statPeriod、sendCount}"),
        dict(h="日志配套与统计", i="查询tasStaObj统计结果", j="EQ", w="tasStaObj统计表", y="tasStaObj统计表{statPeriod、logType|sendCount、recvCount}"),
        dict(g="台账数据可视化展示", h="台账数据可视化展示", i="tasStaObj统计结果图表展示", j="EO", w="tasStaObj统计表", y="tasStaObj统计表{statPeriod|logType、sendCount、recvCount}"),
        dict(g="台账数据上报管理", h="台账数据上报监控", i="统计待上报与失败批次数量", j="EO", w="台账上报内容表、台账上报状态监控表", y="台账上报内容表{reportStatus|batchID}\n台账上报状态监控表{|pendingCount、failCount}"),
        dict(g="台账数据可视化展示", h="logType分布统计", i="按logType导出台账统计报表", j="EO", w="logType台账数量统计表", y="logType台账数量统计表{statPeriod|导出文件路径}"),
        dict(h="上报趋势统计", i="按日统计上报成功失败次数", j="EO", w="台账上报趋势统计表、台账日志上报历史表", y="台账日志上报历史表{reportTime、reportStatus|historyID}\n台账上报趋势统计表{|successCount、failCount}"),
        dict(g="系统漏洞排查日志", h="排查任务日志1020", i="按tskId查询排查任务台账日志", j="EQ", w="系统漏洞排查弱口令扫描日志表", y="系统漏洞排查弱口令扫描日志表{tskId|logID、logType、tskProgress、vulNum}"),
        dict(g="系统漏洞修复日志", h="修复工单日志1050", i="按tktId查询修复工单台账日志", j="EQ", w="系统漏洞修复日志表", y="系统漏洞修复日志表{tktId|logID、logType、method、remedTime}"),
        dict(g="产品漏洞预警日志", h="预警碰撞日志1080", i="按tskId查询预警碰撞台账日志", j="EQ", w="产品漏洞预警日志表", y="产品漏洞预警日志表{tskId|logID、vulNum、vulAstNum、vulInfoNum}"),
        dict(g="系统漏洞利用日志", h="利用过程日志1040", i="按vulInfoID查询利用过程日志", j="EQ", w="系统漏洞利用日志表", y="系统漏洞利用日志表{vulInfoID|logID、logType、engHash、opRslt}"),
        dict(g="漏管平台系统日志", h="平台维护日志1010", i="按loginAccount查询维护操作日志", j="EQ", w="漏管平台维护日志表", y="漏管平台维护日志表{loginAccount、timeStamp|logID、operAction、operDesc}"),
        dict(h="系统运行日志1000", i="按syncStatus查询时钟同步日志", j="EQ", w="漏管平台系统日志表", y="漏管平台系统日志表{syncStatus、timeStamp|logID、logType、syncSource}"),
        dict(g="台账日志基础能力", h="日志采集写入", i="标记日志批次上报成功状态", j="EI", w="台账上报内容表", y="台账上报内容表{batchID、reportStatus|reportStatus、reportTime}"),
        dict(h="日志采集写入", i="标记日志批次上报失败状态", j="EI", w="台账上报内容表", y="台账上报内容表{batchID、failReason|reportStatus、failReason}"),
        dict(h="日志配套与统计", i="关联dataFileID与logInfo记录", j="EI", w="漏管台账日志主表、日志配套文件元数据表", y="漏管台账日志主表{logID、dataFileID|dataFileID}\n日志配套文件元数据表{dataFileID|orderID、fileKey}"),
        dict(g="系统漏洞排查日志", h="排查任务日志1020", i="下载target配套键对象文件", j="EO", w="日志配套文件元数据表", y="日志配套文件元数据表{dataFileID|filePath、fileKey}"),
        dict(g="系统漏洞修复日志", h="修复工单日志1050", i="按remedTime区间查询修复日志", j="EQ", w="系统漏洞修复日志表", y="系统漏洞修复日志表{remedTime|logID、tktName、method}"),
        dict(g="产品漏洞预警日志", h="预警碰撞日志1080", i="导出碰撞分析结果明细", j="EO", w="产品漏洞碰撞分析结果表", y="产品漏洞碰撞分析结果表{logID|导出文件路径}"),
        dict(g="台账数据可视化展示", h="台账数据可视化展示", i="各logType上报成功率对比图", j="EO", w="tasStaObj统计表、台账上报趋势统计表", y="tasStaObj统计表{logType|sendCount}\n台账上报趋势统计表{|successCount、failCount}"),
        dict(g="台账日志基础能力", h="配置管理", i="导入logType码表配置", j="EI", w="logType码表配置表", y="logType码表配置表{logType、logTypeName、logLvlRule|enabled}"),
        dict(h="配置管理", i="导出logType码表配置", j="EO", w="logType码表配置表", y="logType码表配置表{logType|导出文件路径}"),
        dict(g="台账日志基础能力", h="部侧接口对接", i="校验上报批次numLogs与明细一致", j="EQ", w="台账上报内容表、漏管台账日志主表", y="台账上报内容表{batchID、numLogs|numLogs}\n漏管台账日志主表{batchID|logID}"),
    ]

    expected = END - START + 1
    if len(rows) != expected:
        raise ValueError(f"Row count {len(rows)} != {expected}")

    cur_g = cur_h = None
    for row in rows:
        if row.get("g"):
            cur_g = row["g"]
            if row.get("h"):
                cur_h = row["h"]
            else:
                cur_h = None
        elif cur_g:
            row["g"] = cur_g
        if row.get("h"):
            cur_h = row["h"]
        elif cur_h:
            row["h"] = cur_h

    return rows


def unmerge_range(ws, start_row, end_row):
    to_remove = []
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= start_row and merged.max_row <= end_row:
            to_remove.append(str(merged))
    for ref in to_remove:
        ws.unmerge_cells(ref)


def merge_column_blocks(ws, col, start_row, end_row):
    block_start = start_row
    prev = None
    for r in range(start_row, end_row + 2):
        val = ws.cell(r, col).value if r <= end_row else None
        if val != prev:
            if prev is not None and r - 1 > block_start:
                ws.merge_cells(
                    start_row=block_start,
                    start_column=col,
                    end_row=r - 1,
                    end_column=col,
                )
            block_start = r
            prev = val
        elif val is None and prev is not None and r <= end_row:
            if r - 1 > block_start:
                ws.merge_cells(
                    start_row=block_start,
                    start_column=col,
                    end_row=r - 1,
                    end_column=col,
                )
            block_start = r
            prev = None


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["功能点估算"]
    unmerge_range(ws, START, END)
    formula_tpl = {c: ws[f"{c}727"].value for c in FORMULA_COLS}
    rows = build_rows()

    def clear_row(r):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            cell.value = None

    def set_formulas(r):
        for c in FORMULA_COLS:
            val = formula_tpl[c]
            if isinstance(val, str) and val.startswith("="):
                ws[f"{c}{r}"] = val.replace("727", str(r))

    def write_row(r, **data):
        clear_row(r)
        set_formulas(r)
        ws.cell(r, 3).value = 2025
        ws.cell(r, 4).value = "张明峰"
        for key, col in [("f", 6), ("g", 7), ("h", 8), ("i", 9), ("j", 10)]:
            if key in data and data[key] is not None:
                ws.cell(r, col).value = data[key]
        ws.cell(r, 18).value = data.get("r_reuse", 1)
        ws.cell(r, 20).value = data.get("t", 1)
        for key, col in [("w", 23), ("x", 24), ("y", 25), ("z", 26), ("ab", 28)]:
            if key in data and data[key] is not None:
                ws.cell(r, col).value = data[key]

    for idx, data in enumerate(rows):
        write_row(START + idx, **data)

    # Fill down G/H for merge blocks (only top-left cell was written)
    cur_g = cur_h = None
    for r in range(START, END + 1):
        g = ws.cell(r, 7).value
        h = ws.cell(r, 8).value
        if g and g != cur_g:
            cur_g = g
            cur_h = None
        elif cur_g and not g:
            ws.cell(r, 7).value = cur_g
        if h:
            cur_h = h
        elif cur_h:
            ws.cell(r, 8).value = cur_h

    ws.merge_cells(start_row=START, start_column=6, end_row=END, end_column=6)
    merge_column_blocks(ws, 7, START, END)
    merge_column_blocks(ws, 8, START, END)

    wb.save(SRC)
    print(f"Updated {SRC}: rows {START}-{END}, {len(rows)} function points")


if __name__ == "__main__":
    main()
